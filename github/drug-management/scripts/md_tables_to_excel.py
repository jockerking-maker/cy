# -*- coding: utf-8 -*-
"""Convert 论文数据库表_4.3.2.md tables to Excel."""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MD_PATH = Path(__file__).resolve().parent.parent / "论文数据库表_4.3.2.md"
OUT_PATH = Path(__file__).resolve().parent.parent / "论文数据库表_4.3.2.xlsx"

HEADERS = ["字段名称", "类型", "长度", "字段说明", "主键", "默认值"]
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def parse_md_tables(text: str) -> list[tuple[str, list[list[str]]]]:
    tables = []
    blocks = re.split(r"\n(?=表4-\d+)", text)
    for block in blocks:
        block = block.strip()
        if not block.startswith("表4-"):
            continue
        title_line, _, rest = block.partition("\n")
        title = title_line.strip()
        rows = []
        for line in rest.splitlines():
            line = line.strip()
            if not line.startswith("|"):
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if not cells or cells[0] == "字段名称" or set(cells) <= {"---"}:
                continue
            rows.append(cells)
        if rows:
            tables.append((title, rows))
    return tables


def sheet_name(title: str, index: int) -> str:
    match = re.search(r"表4-\d+", title)
    prefix = match.group(0) if match else f"表{index}"
    name_match = re.search(r"（([^）]+)）", title)
    suffix = name_match.group(1) if name_match else str(index)
    name = f"{prefix}_{suffix}"
    return name[:31]


def style_table(ws, start_row: int, title: str, rows: list[list[str]]):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(HEADERS))
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = CENTER

    header_row = start_row + 1
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx in range(len(HEADERS)):
            value = row[col_idx] if col_idx < len(row) else ""
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.border = BORDER
            cell.alignment = CENTER if col_idx != 3 else LEFT

    widths = [16, 10, 8, 42, 8, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return header_row + len(rows) + 2


def build_workbook(tables: list[tuple[str, list[list[str]]]]) -> Workbook:
    wb = Workbook()

    index_ws = wb.active
    index_ws.title = "目录"
    index_ws.append(["序号", "表编号", "表名称", "工作表"])
    for i, (title, _) in enumerate(tables, start=1):
        no_match = re.search(r"表4-\d+", title)
        name_match = re.search(r"表4-\d+\s+(.+?)（", title)
        index_ws.append([i, no_match.group(0) if no_match else "", name_match.group(1) if name_match else title, sheet_name(title, i)])
    for col in range(1, 5):
        index_ws.column_dimensions[get_column_letter(col)].width = [8, 12, 28, 24][col - 1]

    summary_ws = wb.create_sheet("全部表汇总")
    current_row = 1
    for i, (title, rows) in enumerate(tables, start=1):
        current_row = style_table(summary_ws, current_row, title, rows)

    for i, (title, rows) in enumerate(tables, start=1):
        ws = wb.create_sheet(sheet_name(title, i))
        style_table(ws, 1, title, rows)

    return wb


def main():
    text = MD_PATH.read_text(encoding="utf-8")
    tables = parse_md_tables(text)
    if not tables:
        raise RuntimeError("No tables found in markdown file")
    wb = build_workbook(tables)
    wb.save(OUT_PATH)
    print(f"Generated {len(tables)} tables -> {OUT_PATH}")


if __name__ == "__main__":
    main()
